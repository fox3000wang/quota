#!/usr/bin/python3
import os

from openpyxl import Workbook
from openpyxl import load_workbook

# 遍历文件
def findAllFile(base):
    for root, ds, fs in os.walk(base):
        for f in fs:
            if f.endswith('.xlsx'):
                fullname = os.path.join(root, f)
                yield fullname

# 解析xlsx
def parseXlsx(path):
    # 读取
    #sheet = pd.read_excel(path, sheet_name='Sheet1', usecols='B:B')

    workbook = load_workbook(filename=path)
    sheet = workbook["Sheet1"]
    print(sheet['B1'])

# 生成报告
def generateReport():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 42
    wb.save("report.xlsx")

# 主函数
def main():
    base = '.'
    for i in findAllFile(base):
        parseXlsx(i)
    generateReport()    


if __name__ == '__main__':
    main()
